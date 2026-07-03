"""Excel (.xlsx) access in Google Drive via a Service Account.

The job list is kept as a real Excel file in Drive (not converted to native
Google Sheets). We download the file with the Drive API, read/edit it with
openpyxl, and re-upload it over the same file id. The assigned CV number is
written back to column N.
"""
from __future__ import annotations

import datetime
import io
import string
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

from .config import SERVICE_ACCOUNT_FILE, settings

# Drive scope is required to read AND overwrite a user-owned file that was
# shared with the service account.
SCOPES = ["https://www.googleapis.com/auth/drive"]

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Column letters we care about (A..P). Index in this list maps to a row value.
COLUMNS = list(string.ascii_uppercase[:16])  # A..P

# Friendly aliases for the columns we use directly.
COL_BASVURU = "B"
COL_CITY = "F"       # Yer
COL_WORK_MODE = "G"  # Uzaktan: Hybrit / On-site / Remote
COL_LINK = "K"
COL_POZISYON = "L"
COL_ILAN_NO = "M"
COL_CV_NO = "N"
COL_MATCH_RATE = "P"  # post-optimization ATS Match Rate

# 1-based column indexes for openpyxl.
CV_NO_COL_INDEX = COLUMNS.index(COL_CV_NO) + 1        # = 14
MATCH_RATE_COL_INDEX = COLUMNS.index(COL_MATCH_RATE) + 1  # = 16

# Last data row to apply dropdowns to (the sheet has a fixed 1000-row grid).
DROPDOWN_LAST_ROW = 1000

# Dropdowns (data validations) the user maintains on these columns. openpyxl drops
# data validations when it re-saves the file, so we re-apply them on every write to
# keep the user's manual-entry dropdowns intact.
DROPDOWNS: dict[str, list[str]] = {
    "B": ["Geçmiş", "Vazgeçildi", "✓", "+", "++"],   # Başvuru
    "C": ["Var", "Yok"],                              # Easy Apply
    "G": ["Hybrit", "On-site", "Remote"],            # Uzaktan
    "H": ["Full-Time", "Part-Time", "Contract"],     # Çalışma Şekli
}


def _apply_dropdowns(ws) -> None:
    """(Re)create the user's column dropdowns so they survive the openpyxl round-trip."""
    ws.data_validations.dataValidation.clear()
    for col, options in DROPDOWNS.items():
        # Inline list validation: values joined by commas inside double quotes.
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.add(f"{col}2:{col}{DROPDOWN_LAST_ROW}")
        ws.add_data_validation(dv)


@dataclass
class Row:
    """A single spreadsheet row (1-based sheet row number)."""

    number: int  # actual sheet row number (header is row 1, data starts at 2)
    cells: dict[str, str] = field(default_factory=dict)  # column letter -> value

    def get(self, col: str) -> str:
        return self.cells.get(col, "").strip()

    @property
    def link(self) -> str:
        return self.get(COL_LINK)

    @property
    def basvuru(self) -> str:
        return self.get(COL_BASVURU)

    @property
    def city(self) -> str:
        return self.get(COL_CITY)

    @property
    def work_mode(self) -> str:
        return self.get(COL_WORK_MODE)

    @property
    def cv_no(self) -> str:
        return self.get(COL_CV_NO)

    @property
    def ilan_no(self) -> str:
        return self.get(COL_ILAN_NO)

    def key(self) -> str:
        """Stable identifier for de-duplication across runs.

        Based on row number + the link without its volatile query string. We avoid
        column M (İlan Numarası) because it is a spreadsheet formula. The primary
        guard against re-processing is the CV No written to column N anyway.
        """
        base_link = self.link.split("?", 1)[0]
        return f"row:{self.number}:{base_link}"


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value)


def _link_font(cell) -> Font:
    """Hyperlink style (blue, underlined) at 10pt, keeping the cell's font family."""
    return Font(name=cell.font.name, size=10, color="1155CC", underline="single")


class SheetsClient:
    """Drive-backed reader/writer for the .xlsx job list."""

    def __init__(self) -> None:
        creds = Credentials.from_service_account_file(
            str(SERVICE_ACCOUNT_FILE), scopes=SCOPES
        )
        self._drive = build("drive", "v3", credentials=creds, cache_discovery=False)
        self._file_id = settings.spreadsheet_id

    # --- Drive download / upload -------------------------------------------------

    def _download_bytes(self) -> bytes:
        request = self._drive.files().get_media(fileId=self._file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        return buf.read()

    def _load_workbook(self):
        data = self._download_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(data))
        if settings.sheet_name in wb.sheetnames:
            ws = wb[settings.sheet_name]
        else:
            ws = wb.active
        return wb, ws

    def _upload(self, wb) -> None:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        media = MediaIoBaseUpload(buf, mimetype=XLSX_MIME, resumable=True)
        self._drive.files().update(fileId=self._file_id, media_body=media).execute()

    # --- Public API (same surface as before) -------------------------------------

    def get_rows(self) -> list[Row]:
        """Return all non-empty data rows (from sheet row 2 onward)."""
        _, ws = self._load_workbook()
        rows: list[Row] = []
        for r in range(2, ws.max_row + 1):
            cells: dict[str, str] = {}
            non_empty = False
            for j, col in enumerate(COLUMNS):
                val = _cell_to_str(ws.cell(row=r, column=j + 1).value)
                cells[col] = val
                if val.strip():
                    non_empty = True
            if non_empty:
                rows.append(Row(number=r, cells=cells))
        return rows

    @staticmethod
    def _as_int(val: str) -> int | None:
        val = val.strip()
        if not val:
            return None
        try:
            return int(float(val))  # tolerates "200" and "200.0"
        except ValueError:
            return None

    def next_cv_number(self, rows: list[Row]) -> int:
        """Highest numeric value in column N (CV No) + 1, minimum 1."""
        nums = [n for r in rows if (n := self._as_int(r.cv_no)) is not None]
        return (max(nums) + 1) if nums else 1

    def write_cv_number(self, row_number: int, cv_no: int) -> None:
        """Re-download the latest file, set N{row}, and re-upload.

        Downloading fresh right before writing keeps the clobber window tiny so
        we don't overwrite edits you made elsewhere in the file.
        """
        wb, ws = self._load_workbook()
        ws.cell(row=row_number, column=CV_NO_COL_INDEX).value = cv_no
        _apply_dropdowns(ws)
        self._upload(wb)

    def write_match_rate(self, row_number: int, match_rate: float) -> None:
        """Write the post-optimization ATS Match Rate to column P."""
        wb, ws = self._load_workbook()
        ws.cell(row=row_number, column=MATCH_RATE_COL_INDEX).value = match_rate
        _apply_dropdowns(ws)
        self._upload(wb)

    def write_cv_and_match(
        self, row_number: int, cv_no: int, match_rate: float | None = None
    ) -> None:
        """Write CV No (N) and, if available, Match Rate (P) in a single upload.

        Doing both in one download-edit-upload minimizes how often (and how long) the
        file is replaced, reducing the chance of clobbering edits you make in the
        browser at the same time.
        """
        wb, ws = self._load_workbook()
        ws.cell(row=row_number, column=CV_NO_COL_INDEX).value = cv_no
        if match_rate is not None:
            ws.cell(row=row_number, column=MATCH_RATE_COL_INDEX).value = match_rate
        _apply_dropdowns(ws)
        self._upload(wb)

    # --- New-row append (Huntr import) -------------------------------------------

    @staticmethod
    def _col(letter: str) -> int:
        return COLUMNS.index(letter) + 1

    def _last_data_row(self, ws) -> int:
        """Highest row that has data in any key column (A/J/K/L/N)."""
        key_cols = [self._col(c) for c in ("A", "J", "K", "L", "N")]
        last = 1
        for r in range(2, ws.max_row + 1):
            if any(_cell_to_str(ws.cell(row=r, column=c).value).strip() for c in key_cols):
                last = r
        return last

    def _set_link_cell(self, cell, url: str) -> None:
        """Write a clickable hyperlink cell (blue, underlined)."""
        cell.value = url
        if url:
            cell.hyperlink = url
            cell.font = _link_font(cell)

    def append_job_rows(self, jobs: list[dict], cv_no_marker: str | None = None) -> list[int]:
        """Append rows for new jobs — writes K (link) + E (date) + A (No) + I + M.

        Each job dict needs at least {"url": ...}. It MAY also carry Huntr-sourced
        info ("company"/"title"/"city") which is written to J/L/F right away — used by
        the info-only (CV_GENERATION off) path so the sheet is filled straight from
        Huntr instead of re-scraping the (often expired) job page. When cv_no_marker is
        given it is written to N (CV No) so the row counts as handled with no CV.

        In CV-generation mode the caller passes url-only dicts and no marker, so the
        remaining columns (C/F/H/J/L) are filled later by write_processed_row from the
        scraped page — exactly as before. Single download-edit-upload.
        """
        if not jobs:
            return []
        wb, ws = self._load_workbook()
        last = self._last_data_row(ws)
        date_fmt = ws.cell(row=last, column=self._col("E")).number_format if last >= 2 else None
        # Continue the "No" (A) numbering from the most recent numeric value above.
        last_a = None
        for rr in range(last, 1, -1):
            v = self._as_int(_cell_to_str(ws.cell(row=rr, column=self._col("A")).value))
            if v is not None:
                last_a = v
                break
        today = datetime.datetime(*datetime.date.today().timetuple()[:3])

        written: list[int] = []
        r = last + 1
        for job in jobs:
            if last_a is not None:
                last_a += 1
                ws.cell(row=r, column=self._col("A")).value = last_a
            e_cell = ws.cell(row=r, column=self._col("E"))
            e_cell.value = today
            if date_fmt:
                e_cell.number_format = date_fmt
            ws.cell(row=r, column=self._col("I")).value = "İş"
            self._set_link_cell(ws.cell(row=r, column=self._col("K")), job.get("url", ""))
            # İlan No formula: extract the id between "/jobs/view/" and the next "/".
            ws.cell(row=r, column=self._col("M")).value = (
                f'=IFERROR(MID(K{r},FIND("/jobs/view/",K{r})+11,'
                f'FIND("/",K{r}&"/",FIND("/jobs/view/",K{r})+11)-FIND("/jobs/view/",K{r})-11),"")'
            )
            # Optional Huntr-sourced info (info-only path); only set when provided.
            if job.get("company"):
                ws.cell(row=r, column=self._col("J")).value = job["company"]
            if job.get("title"):
                ws.cell(row=r, column=self._col("L")).value = job["title"]
            if job.get("city"):
                ws.cell(row=r, column=self._col("F")).value = job["city"]
            if cv_no_marker is not None:
                ws.cell(row=r, column=CV_NO_COL_INDEX).value = cv_no_marker
            written.append(r)
            r += 1

        _apply_dropdowns(ws)
        self._upload(wb)
        return written

    def _fill_page_fields(self, ws, row_number: int, fields: dict) -> None:
        """Fill ONLY-EMPTY C/F/H/J/L from the scraped page (J as a company hyperlink).
        G is never touched (filled manually). Does not upload."""

        def fill(col: str, val: str):
            if not val:
                return None
            cell = ws.cell(row=row_number, column=self._col(col))
            if _cell_to_str(cell.value).strip() == "":
                cell.value = val
                return cell
            return None

        fill("C", fields.get("C", ""))   # Easy Apply (only set for non-LinkedIn = "Yok")
        fill("F", fields.get("F", ""))   # city
        fill("H", fields.get("H", ""))   # employment type
        fill("L", fields.get("L", ""))   # title
        j_cell = fill("J", fields.get("J", ""))  # company
        if j_cell is not None and fields.get("J_url"):
            j_cell.hyperlink = fields["J_url"]
            j_cell.font = _link_font(j_cell)

    def write_processed_row(
        self, row_number: int, fields: dict, cv_no: int, match_rate: float | None = None
    ) -> None:
        """Fill ONLY-EMPTY C/F/H/J/L from the scraped page (J as a company hyperlink),
        plus N (CV No) and P (Match Rate). G is never touched (filled manually).
        Single download-edit-upload."""
        wb, ws = self._load_workbook()
        self._fill_page_fields(ws, row_number, fields)
        ws.cell(row=row_number, column=CV_NO_COL_INDEX).value = cv_no
        if match_rate is not None:
            ws.cell(row=row_number, column=MATCH_RATE_COL_INDEX).value = match_rate
        _apply_dropdowns(ws)
        self._upload(wb)

    def write_info_only_row(self, row_number: int, fields: dict, cv_no_marker: str) -> None:
        """CV-generation-OFF path: fill ONLY-EMPTY C/F/H/J/L from the scraped page and
        write a non-numeric marker (e.g. "Yok") to N so the row counts as handled but is
        never assigned a CV. No Match Rate is written. Single download-edit-upload."""
        wb, ws = self._load_workbook()
        self._fill_page_fields(ws, row_number, fields)
        ws.cell(row=row_number, column=CV_NO_COL_INDEX).value = cv_no_marker
        _apply_dropdowns(ws)
        self._upload(wb)
